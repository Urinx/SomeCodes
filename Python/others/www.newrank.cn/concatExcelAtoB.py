#! /usr/bin/env python
from openpyxl import load_workbook
import xlrd

[A, B] = sys.argv[1:3]

wbB = load_workbook(B)
wsB = wbB.get_sheet_by_name(wbB.get_sheet_names()[0])

wbA = xlrd.open_workbook(A)
wsA = wbA.sheets()[0]
for r in range(1, wsA.nrows):
	wsB.append([wsA.cell(r, i).value for i in range(wsB.columns)])

wbB.save(B)