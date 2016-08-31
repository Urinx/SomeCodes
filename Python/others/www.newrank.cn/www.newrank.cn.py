#!/usr/bin/env python
# coding: utf-8
import requests
import hashlib
import random
import json
import re
import getopt, sys
import os
from multiprocessing.dummy import Pool as ThreadPool 

def md5(s):
	m = hashlib.md5()
	m.update(s)
	return m.hexdigest()

# 随机9位16进制数字
def genNonce():
	return ''.join([hex(int(16 * random.random()))[-1] for _ in range(9)])

def Log(dict):
	for (key, value) in dict.items():
		if type(value) == list:
			print key,':','[',', '.join(value),']'
		else: print key,':',value
	print '='*30

# keyName -> 地名: 武汉, 贵阳
# return datalist
def searchWeixinData(keyName):
	host = 'http://www.newrank.cn'
	baseUrl = '/xdnphb/data/weixinuser/searchWeixinData'
	nonce = genNonce()
	h = baseUrl + '?AppKey=joker&hasDeal=false&isBind=false&isNewRank=true&isTag=false&keyName={0}&nonce={1}'.format(keyName, nonce)

	data = {
		'hasDeal': 'false',
		'isBind': 'false',
		'isNewRank': 'true',
		'isTag': 'false',
		'keyName': keyName,
		'nonce': nonce,
		'xyz': md5(h),
	}

	r = requests.post(host + baseUrl, data = data)
	result = json.loads(r.content)
	list = json.loads(result['value']['list'])['result']
	return list

def fetchWeixinFansCounts(account):
	fans_counts = ''
	url = 'http://www.newrank.cn/public/info/detail.html?account=' + account
	r = requests.get(url)
	m = re.search(r'<div class="detail-fans-counts">\s+(.+)\s+</div>', r.content)
	if m: fans_counts = m.group(1)
	return fans_counts

def bindAccountState(uuid):
	host = 'http://www.newrank.cn'
	baseUrl = '/xdnphb/data/bindAccountState'
	nonce = genNonce()
	h = baseUrl + '?AppKey=joker&uuid={0}&nonce={1}'.format(uuid, nonce)

	data = {
		'uuid': uuid,
		'nonce': nonce,
		'xyz': md5(h),
	}

	r = requests.post(host + baseUrl, data = data)
	state_code = int(json.loads(r.content)['value'])
	return ['未知', '询购', '未被认领', '已认领', '马上认领', '已认领', '马上认领'][state_code]

# return list:
# {
# 	'name': // 公众号名称
#	'account': // 微信号
#	'description': // 公众号介绍
#	'weekLog1pmark': // 新榜指数
#	'type': // 分类
#	'tags': // 标签（每个标签一个字段）
#	'fans_counts': // 预估粉丝数
#	'state': // 认领状态
# }
def getTop5WeixinData(keyName):
	top5 = searchWeixinData(keyName)[:5]
	newTop5 = []

	def triphtml(s):
		return s.replace("<font  color='red'>","").replace("</font>","").replace('<span title="','').replace('">--/--</span>','')

	for i in top5:
		description = triphtml(i['description']) if i.has_key('description') else ''
		pmark = round(float(i['weekLog1pmark']),1) if i.has_key('weekLog1pmark') else ''
		newTop5.append({
			'name': triphtml(i['name']),
			'account': i['account'],
			'description': description,
			'weekLog1pmark': pmark,
			'type': i['type'],
			'tags': ','.join([triphtml(t) for t in i['tags']]),
			'fans_counts': triphtml(fetchWeixinFansCounts(i['account'])),
			'state': bindAccountState(i['uuid']),
			'area': keyName
		})

	return newTop5

LOG_MODE = False
SAVE_DATA = False
TEMP_FILE = 'data.xlsx'
WEIXIN_DATA = []
START_INDEX = 1
END_INDEX = 0

def process(city):
	result = getTop5WeixinData(city)
	print 'Top 5 in '+city
	print '='*30

	if LOG_MODE:
		for i in result: Log(i)

	return result

def usage():
	print 'usage:'
	print sys.argv[0] + ' -v -k [key word] -o [output file]'
	print sys.argv[0] + ' -v --txt=[word list file]'
	print sys.argv[0] + ' -v --excel=[excel file]'
	print sys.argv[0] + ' -v -s 1 -e 10 --excel=[excel file] -o [output file]'
	print sys.argv[0] + ' -h'

def main(argv):
	global SAVE_DATA, LOG_MODE, WEIXIN_DATA, START_INDEX, END_INDEX
	opts, args = getopt.getopt(argv[1:], "hvk:o:s:e:", ["help", "excel=", "txt="])
	for op, value in opts:
		if op in ('-h', '--help'): usage()
		elif op in ('-s'): START_INDEX = int(value)
		elif op in ('-e'): END_INDEX = int(value)
		elif op in ('-k'):
			process(value)
		elif op in ('-v'):
			LOG_MODE = True
		elif op in ('--txt'):
			cityList = []
			with open(value, 'r') as f:
				cityList = [line.strip() for line in f.readlines()]
			pool = ThreadPool()
			result = pool.map(process, cityList)
			pool.close()
			pool.join()
			for r in result: WEIXIN_DATA.extend(r)
		elif op in ('--excel'):
			import xlrd
			wb = xlrd.open_workbook(value)
			ws = wb.sheets()[0]

			cityList = []
			lastCity = ''
			if END_INDEX == 0: END_INDEX = ws.nrows
			for r in range(START_INDEX, END_INDEX):
				c = ws.cell(r, 3).value
				if len(c) > 2:
					if c[-1] == u"市":
						c = c[:-1]
						lastCity = c
					elif c[-1] == u"旗": pass
					elif c[:-1] == lastCity: pass
					else: c = c[:-1]
				cityList.append(c.encode('utf-8'))

			pool = ThreadPool()
			result = pool.map(process, cityList)
			pool.close()
			pool.join()
			for r in result: WEIXIN_DATA.extend(r)

		elif op in ('-o'):
			from openpyxl import Workbook
			from openpyxl import load_workbook
			wb = None
			ws = None
			col = ['name','account','area','fans_counts','weekLog1pmark','state','type','tags','description']
			if os.path.isfile(value):
				wb = load_workbook(value)
				ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
			else:
				wb = Workbook()
				ws = wb.active
				ws.append(col)
			for i in WEIXIN_DATA:
				ws.append([i[c] for c in col])
			wb.save(value)

if __name__ == '__main__':
	main(sys.argv)
