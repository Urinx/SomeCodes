#!/usr/bin/env python
# -*- coding: UTF-8 -*-

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import sys
from openpyxl import load_workbook
from time import sleep
import ConfigParser

cp = ConfigParser.ConfigParser()
cp.read('mail.conf')
MAIL_HOST = cp.get('account','host')
MAIL_USER = cp.get('account','user')
MAIL_PASS = cp.get('account','password')
MAIL_FROMNAME = cp.get('mail','from_name')
MAIL_SUBJECT = cp.get('mail','subject')
MAIL_LIST = cp.get('mail', 'list')
MAIL_CONTENT = cp.get('mail', 'content')
MAIL_ATTACH = cp.get('mail', 'attach')
PAUSE_TIME = cp.getint('mail','pause_time')
START_INDEX = cp.getint('mail', 'index')
COL_NAME = cp.getint('excel', 'name_col')
COL_MAIL = cp.getint('excel', 'mail_col')
COL_RESULT = cp.getint('excel', 'result_col')
STMP_SERVER = smtplib.SMTP_SSL(MAIL_HOST)

def sendTextMail(fromName, to, subject, content):
	message = MIMEText(content, 'plain', 'utf-8')
	message['From'] = Header(fromName, 'utf-8')
	message['To'] =  Header(to)
	message['Subject'] = Header(subject, 'utf-8')
	STMP_SERVER.sendmail(MAIL_USER, [to], message.as_string())

def sendHTMLMail(fromName, to, subject, content):
	message = MIMEText(content, 'html', 'utf-8')
	message['From'] = Header(fromName, 'utf-8')
	message['To'] =  Header(to)
	message['Subject'] = Header(subject, 'utf-8')
	STMP_SERVER.sendmail(MAIL_USER, [to], message.as_string())

def sendAttachMail(fromName, to, subject, content, file):
	message = MIMEMultipart()
	message['From'] = Header(fromName, 'utf-8')
	message['To'] =  Header(to)
	message['Subject'] = Header(subject, 'utf-8')
	message.attach(MIMEText(content, 'html', 'utf-8'))

	att = MIMEText(open(file, 'rb').read(), 'base64', 'utf-8')
	att["Content-Type"] = 'application/octet-stream'
	att["Content-Disposition"] = 'attachment; filename="{0}"'.format(file)
	message.attach(att)

	STMP_SERVER.sendmail(MAIL_USER, [to], message.as_string())

def usage():
	print '帮助:'
	print '  python qqmail.py [邮箱帐号] [邮箱密码] [邮箱列表文件] [邮件内容文件] [附件] [从第几个开始]'


if __name__ == '__main__':
	content = open(MAIL_CONTENT,'r').read()
	wb = load_workbook(MAIL_LIST)
	ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])

	print '邮箱服务器:',MAIL_HOST
	print '登录帐号:',MAIL_USER
	print '登录密码: *********'
	STMP_SERVER.login(MAIL_USER, MAIL_PASS)
	print '登录邮箱服务器 [成功]'
	print '====================='
	print '邮件内容文件:',MAIL_CONTENT
	print '邮箱列表文件:',MAIL_LIST
	print '邮件附件:',MAIL_ATTACH
	print '====================='
	print '共有{0}位联系人, 现在从第{1}位开始发送邮件'.format(ws.max_row - 1, START_INDEX)

	for i in xrange(START_INDEX+1, ws.max_row+1):
		name = ws.cell(row=i,column=COL_NAME).value.encode('utf-8')
		mail = ws.cell(row=i,column=COL_MAIL).value
		html = content.replace('#NAME#',name)
		try:
			sendAttachMail(MAIL_FROMNAME, mail, MAIL_SUBJECT, html, MAIL_ATTACH)
			print '[{0}] 发送邮件 -> {1}: {2} [成功]'.format(i-1, name, mail)
			ws.cell(row=i,column=COL_RESULT).value = '已发送'
		except:
			print '[{0}] 发送邮件 -> {1}: {2} [失败]'.format(i-1, name, mail)
			ws.cell(row=i,column=COL_RESULT).value = '发送失败'
		wb.save(MAIL_LIST)
		sleep(PAUSE_TIME)

	print '====================='
	print '断开连接'
	STMP_SERVER.quit()
		
		
		




